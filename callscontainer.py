import time
from openpyxl import Workbook

from magicDB import *

from myExceptions import *

__version__ = "1.5.3"


###############################################################################################################################################################################
###############################################################################################################################################################################
###############################################################################################################################################################################

class Call(object):
    """
    This class holds data about one call
    """
    def __init__(self, callId=None, creator=None, totalDuration=None, answerDuration=None, typeId=None, callType=None, client=None,
                 product=None, device=None, statusId=None, status=None, supportId=None, support=None, description=None, inPlan=None,
                 openingDate=None, answeringDate=None, endDate=None, detractor=None):
        
        self.callId = int(callId)
        self.creator = creator
        self.totalDuration = totalDuration
        self.answerDuration = answerDuration

        if totalDuration < 0:
            totalDuration = abs(int(totalDuration))
            self.formatedTotalDuration = "-%02d:%02d" % (totalDuration//60, totalDuration%60)
        else:
            totalDuration = abs(int(totalDuration))
            self.formatedTotalDuration = "%02d:%02d" % (totalDuration//60, totalDuration%60)

        if answerDuration is not None:    
            self.formatedAnswerDuration = ("%02d:%02d" % ((int(answerDuration)//60), (int(answerDuration)%60)))
        else:
            self.formatedAnswerDuration = ""

        self.typeId = int(typeId)
        self.callType = callType
        self.client = client
        self.product = product
        self.device = device
        self.statusId = statusId
        self.status = status
        self.supportId = supportId
        self.support = support
        self.description = description
        self.insidePlan = int(inPlan)
        self.openingDate = openingDate
        self.answeringDate = answeringDate
        self.endDate = endDate
        self.detractor = detractor

        dateFormat = "%Y-%m-%d %H:%M:%S"
        strippedOpeningDate = time.strptime(str(openingDate), dateFormat)
        self.formatedOpeningDate = "%02d/%02d/%d %02d:%02d" % (strippedOpeningDate[2], strippedOpeningDate[1], strippedOpeningDate[0], strippedOpeningDate[3], strippedOpeningDate[4])

        self.formatedAnsweringDate = ""
        if answeringDate is not None:
            strippedAnsweringDate = time.strptime(str(answeringDate), dateFormat)
            self.formatedAnsweringDate = "%02d/%02d/%d %02d:%02d" % (strippedAnsweringDate[2], strippedAnsweringDate[1], strippedAnsweringDate[0], strippedAnsweringDate[3], strippedAnsweringDate[4])

        self.formatedEndDate = ""
        if endDate is not None:
            strippedEndDate = time.strptime(str(endDate), dateFormat)
            self.formatedEndDate = "%02d/%02d/%d %02d:%02d" % (strippedEndDate[2], strippedEndDate[1], strippedEndDate[0], strippedEndDate[3], strippedEndDate[4])

###############################################################################################################################################################################
        
    def updateDuration(self):
        """
        This method updates the call duration
        """

        # --> DATABASE CONNECTION OBJECT #
        myDBConnection = DBConnection()
        
        fields = [
            "TIMESTAMPDIFF(MINUTE, cha_data_hora_abertura, NOW()) AS duracao_total",
            "IF(cha_status > 1, TIMESTAMPDIFF(MINUTE, cha_data_hora_atendimento, NOW()), 0) AS duracao_atendimento",
            "cha_data_hora_abertura",
            "cha_data_hora_atendimento",
            "cha_data_hora_termino"
            ]
            
        tables = [
            ("chamados", "", "")
            ]

        where = [
             "cha_id = '" + str(self.callId) + "' "           
             ]
        
        querySuccess, queryResult = myDBConnection.selectQuery(fields, tables, where)
        if not querySuccess:
            raise DatabaseConnectionError("Durante a atualização da duração do Chamado.",
                                          "Falha ao tentar executar a consulta no banco.\nContate o administrador do sistema.")
            return

        
        if len(queryResult) > 0:
            self.totalDuration = int(queryResult[0][0])
            self.answerDuration = int(queryResult[0][1])
            self.openingDate = queryResult[0][2]
            self.answeringDate = queryResult[0][3]
            self.endDate = queryResult[0][4]
        
        if self.totalDuration < 0:
            totalDuration = abs(self.totalDuration)
            self.formatedTotalDuration = "-%02d:%02d" % (totalDuration//60, totalDuration%60)
        else:
            self.formatedTotalDuration = "%02d:%02d" % (self.totalDuration//60, self.totalDuration%60)

        self.formatedAnswerDuration = "%02d:%02d" % (self.answerDuration//60, self.answerDuration%60)

        dateFormat = "%Y-%m-%d %H:%M:%S"
        strippedOpeningDate = time.strptime(str(self.openingDate), dateFormat)
        self.formatedOpeningDate = "%02d/%02d/%d %02d:%02d" % (strippedOpeningDate[2], strippedOpeningDate[1], strippedOpeningDate[0], strippedOpeningDate[3], strippedOpeningDate[4])

        self.formatedAnsweringDate = ""
        if self.answeringDate is not None:
            strippedAnsweringDate = time.strptime(str(self.answeringDate), dateFormat)
            self.formatedAnsweringDate = "%02d/%02d/%d %02d:%02d" % (strippedAnsweringDate[2], strippedAnsweringDate[1], strippedAnsweringDate[0], strippedAnsweringDate[3], strippedAnsweringDate[4])

        self.formatedEndDate = ""
        if self.endDate is not None:
            strippedEndDate = time.strptime(str(self.endDate), dateFormat)
            self.formatedEndDate = "%02d/%02d/%d %02d:%02d" % (strippedEndDate[2], strippedEndDate[1], strippedEndDate[0], strippedEndDate[3], strippedEndDate[4])

        
###############################################################################################################################################################################

    def hasHelpers(self):
        """
        This method verifies if there is any helper in the call. If there is no helpers return false, else returns True and the list of helpers
        """

        # --> DATABASE CONNECTION OBJECT #
        myDBConnection = DBConnection()
        
        fields = [
            "col_nome"
            ]
            
        tables = [
            ("ajudantes_chamados", "", ""),
            ("colaboradores", "LEFT", "ajc_colaborador = col_id")
            ]

        where = [
             "ajc_chamado = '" + str(self.callId) + "' "
             ]
        if self.supportId:
            where.append("ajc_colaborador != '" + str(self.supportId) + "' ")


        querySuccess, queryResult = myDBConnection.selectQuery(fields, tables, where)

        if not querySuccess:
            raise DatabaseConnectionError("Durante a busca pelos ajudantes no chamado.",
                                          "Falha ao tentar executar a consulta no banco.\nContate o administrador do sistema.")
            return

        if len(queryResult) < 1:
            return (False, [])

        callHelpers = []
        for helper in queryResult:
            callHelpers.append(helper[0])

        return(True, callHelpers)


###############################################################################################################################################################################
        
    def isHelping(self, userId):
        """
        This method check if the user passed as Argument is helping in this call
        """

        # --> DATABASE CONNECTION OBJECT #
        myDBConnection = DBConnection()

        fields = [
            "col_nome"
            ]
            
        tables = [
            ("ajudantes_chamados", "", ""),
            ("colaboradores", "LEFT", "ajc_colaborador = col_id")
            ]

        where = [
             "ajc_chamado = '" + str(self.callId) + "' ",
             "ajc_colaborador = '" + str(userId) + "' "
             ]

        querySuccess, queryResult = myDBConnection.selectQuery(fields, tables, where)

        if not querySuccess:
            raise DatabaseConnectionError("Durante a adição de ajudantes ao chamado.",
                                          "Falha ao tentar executar a consulta no banco.\nContate o administrador do sistema.")
            return False

        if len(queryResult) < 1:
            return False

        return True


###############################################################################################################################################################################

    def addHelper(self, userId):
        """
        This method is used to insert a helper to a Call
        """

        # --> DATABASE CONNECTION OBJECT #
        myDBConnection = DBConnection()
        
        table = "ajudantes_chamados"

        fields = ["ajc_chamado", "ajc_colaborador", "ajc_data_hora"]

        values = [("'" + str(self.callId) + "'", "'" + str(userId) + "'", "NOW()")]

        inserted, insertedId = myDBConnection.insertQuery(table, fields, values)

        if not inserted:
            raise DatabaseConnectionError("Durante a adição de ajudantes ao chamado.",
                                          "Falha ao tentar executar a inserção no banco.\nContate o administrador do sistema.")
            return 


###############################################################################################################################################################################
     

    def transferCallFromTo(self, oldUser, newUser):
        """
        This method is used to transfer a call to another user
        """

        # --> DATABASE CONNECTION OBJECT #
        myDBConnection = DBConnection()

        queryList = []
        querySuccess = False

        ##########################################################
        #### --> Finish the call answer for the previous user ####
        ##########################################################

        tables = ["atendimentos_chamados"]

        fieldsAndValues = [("atc_data_hora_termino", "NOW()")]

        conditions = [
            "atc_chamado = '" + str(self.callId) + "'",
            "atc_colaborador = '" + str(oldUser) + "'"
            ]
        

        queryList.append(myDBConnection.updateQuery(tables, fieldsAndValues, conditions, queryBuild=True))

        #########################################################################
        #### --> After that creates a new register for the new anwering user ####
        #########################################################################
        
        table = "atendimentos_chamados"

        fields = ["atc_chamado", "atc_colaborador", "atc_data_hora_inicio"]

        values = [("'" + str(self.callId) + "'", "'" + str(newUser) + "'", "NOW()")]

        queryList.append(myDBConnection.insertQuery(table, fields, values, queryBuild=True))

        querySuccess = myDBConnection.aglutinatedQuery(queryList)
 
        if not querySuccess:
            raise DatabaseConnectionError("Durante a Transferencia do Chamado.",
                                          "Falha ao tentar executar a inserção no banco.\nContate o administrador do sistema.")
            return 

        self.updateCallData()

###############################################################################################################################################################################

    def updateCallData(self):
        """
        This method is used to update the Call Data from Database
        """

        # --> DATABASE CONNECTION OBJECT #
        myDBConnection = DBConnection()
        
        fields = [
                "TIMESTAMPDIFF(MINUTE, cha_data_hora_abertura, NOW()) AS duracao_total",
                "IF(cha_status > 1, TIMESTAMPDIFF(MINUTE, cha_data_hora_atendimento, NOW()), 0) AS duracao_atendimento",
                "cha_status",
                "stc_descricao",
                "atc_colaborador",
                "col_nome",
                "cha_data_hora_abertura",
                "cha_data_hora_atendimento",
                "cha_data_hora_termino"
                ]
            
        tables = [
            ("chamados", "", ""),            
            ("status_chamado", "LEFT", "cha_status = stc_id"),
            ("atendimentos_chamados", "LEFT", "cha_id = atc_chamado"),
            ("colaboradores", "LEFT", "col_id = atc_colaborador")
            ]

        where = [
             "cha_id = '" + str(self.callId)+ "' ",
             "atc_data_hora_termino IS NULL"
             ]

        querySuccess, queryResult = myDBConnection.selectQuery(fields, tables, where)

        if not querySuccess:
            raise DatabaseConnectionError("Durante a atualização dos dados do chamado a ser atendido",
                                          "Falha ao tentar executar a seleção no banco.\nContate o administrador do sistema.")
            return False

        self.totalDuration = queryResult[0][0]
        self.answerDuration = queryResult[0][1]
        self.formatedTotalDuration = ("%02d:%02d" % ((int(queryResult[0][0])//60), (int(queryResult[0][0])%60)))
        self.formatedAnswerDuration = ("%02d:%02d" % ((int(queryResult[0][1])//60), (int(queryResult[0][1])%60)))
        self.statusId = queryResult[0][2]
        self.status = queryResult[0][3]
        self.supportId = queryResult[0][4]
        self.support = queryResult[0][5]
        self.openingDate = queryResult[0][6]
        self.answeringDate = queryResult[0][7]
        self.endDate = queryResult[0][8]

###############################################################################################################################################################################
    def changeCallDateTimes(self, beginningDate, endDate):
        """
        This method is responsible for changing the beginning and the End call Datetime
        """
        # --> DATABASE CONNECTION OBJECT #
        myDBConnection = DBConnection()
        
        tables = ["chamados"]

        fieldsAndValues = [
            ("cha_data_hora_atendimento", "'" + beginningDate + "'"),
            ("cha_data_hora_termino", "'" + endDate + "'")
            ]

        conditions = [
            "cha_id = '" + str(self.callId) + "'"
            ]

        updated, updatedId = myDBConnection.updateQuery(tables, fieldsAndValues, conditions)

        if not updated:
            raise DatabaseConnectionError("Durante a alteração das datas de início e fim do chamado.",
                                          "Falha ao tentar executar a atualização no banco.\nContate o administrador do sistema.")
            return False

        
        return True



###############################################################################################################################################################################
      
    def lockCall(self, lock=True):
        """
        This method is used to set the field "cal_visualizado" to 1 or to 0, informing that the call is being
        visualized by another user
        """

        # --> DATABASE CONNECTION OBJECT #
        myDBConnection = DBConnection()
        
        tables = ["chamados"]

        fieldsAndValues = [("cha_visualizado", ("'1'" if lock else "'0'"))]

        conditions = [
            "cha_id = '" + str(self.callId) + "'"
            ]

        updated, updatedId = myDBConnection.updateQuery(tables, fieldsAndValues, conditions)

        if not updated:
            raise DatabaseConnectionError("Durante o bloqueio do chamado para não poder ser visualizado por outro usuário.",
                                          "Falha ao tentar executar a atualização no banco.\nContate o administrador do sistema.")
            return False

        
        return True


###############################################################################################################################################################################

    def isLockedCall(self):
        """
        This method verifies if a call is being visualized by another user
        """

        # --> DATABASE CONNECTION OBJECT #
        myDBConnection = DBConnection()

        fields = ["cha_visualizado"]

        tables = [("chamados", "", "")]

        where = [
            "cha_id = '" +str(self.callId)+ "'",
            "cha_status = 1"
            ]

        querySuccess, queryResult = myDBConnection.selectQuery(fields, tables, where)

        if not querySuccess:
            raise DatabaseConnectionError("Durante a verificação se o chamado está sendo visualizado por outro.",
                                          "Falha ao tentar executar a consulta no banco.\nContate o administrador do sistema.")
            return

        if len(queryResult) > 0:
            if int(queryResult[0][0]) > 0:
                return True
        
        
        return False


###############################################################################################################################################################################


    def setCallAsBeingAnswered(self, idResponsible):
        """
        This method switch the state of the call, set a responsible for the call and
        updates its answer time
        """

        # --> DATABASE CONNECTION OBJECT #
        myDBConnection = DBConnection()
       
        # --> firstly inserts the user in the responsible database #
        table = "atendimentos_chamados"

        fields = ["atc_chamado", "atc_colaborador", "atc_data_hora_inicio"]

        values = [
            ("'" + str(self.callId) + "'", "'" + str(idResponsible) + "'", "NOW()")
            ]

        inserted, insertedId = myDBConnection.insertQuery(table, fields, values)

        if not inserted:
            raise DatabaseConnectionError("Durante a criação de um novo atendimento do chamado.",
                                          "Falha ao tentar executar a inserção no banco.\nContate o administrador do sistema.")
            return 

        # --> after retrieving the inserted id, update the call status, answer datetime and responsible #
        else:
            tables = ["chamados"]

            fieldsAndValues = [
                ("cha_status", "'2'"),
                ("cha_data_hora_atendimento", "NOW()")            
                ]

            conditions = [
                "cha_id = '" + str(self.callId) + "'"
                ]

            updated, updatedId = myDBConnection.updateQuery(tables, fieldsAndValues, conditions)

            if not updated:
                raise DatabaseConnectionError("Durante a modificação do status do chamado para atendendo.",
                                              "Falha ao tentar executar a atualização no banco.\nContate o administrador do sistema.")
                return


###############################################################################################################################################################################


    def giveUpFromCall(self, idResponsible):
        """
        This method is called by an user to give up from a call
        """

        # --> DATABASE CONNECTION OBJECT #
        myDBConnection = DBConnection()
        
        tables = ["atendimentos_chamados"]
      

        conditions = [
            "atc_chamado = '" +str(self.callId)+ "'"
            ]

        deleted, deletedId = myDBConnection.deleteQuery(tables, conditions)

        if not deleted:
            raise DatabaseConnectionError("Durante a desistência do atendimento do chamado.",
                                          "Falha ao tentar executar a remoção no banco.\nContate o administrador do sistema.")
            return


        if deleted:
            
            tables = ["chamados"]

            fieldsAndValues = [
                ("cha_status", "1"),
                ("cha_data_hora_termino", "NULL"),
                ("cha_visualizado", "'0'")
                ]

            conditions = [
                "cha_id = '" +str(self.callId)+ "'"
                ]

            updated, updatedId = myDBConnection.updateQuery(tables, fieldsAndValues, conditions)

            if not updated:
                raise DatabaseConnectionError("Durante a desistência do atendimento do chamado.",
                                              "Falha ao tentar executar a atualização no banco.\nContate o administrador do sistema.")
                return

###############################################################################################################################################################################
   

    def closeCall(self, detractorId, actionTaked):
        """
        This method is called to finish a call. It inserts the action taked to solve the call, after that update the "atendimentos_chamados" table with the
        finish date of the call and change the call Status to concluded
        """

        # --> DATABASE CONNECTION OBJECT #
        myDBConnection = DBConnection()
        
        queryList = []
        querySuccess = False
        
        ##############################################
        # --> Frirstly updates the responsible table #
        ##############################################
                      
        tables = ["atendimentos_chamados"]
        
        fieldsAndValues = [("atc_data_hora_termino", "NOW()")]
         
        conditions = [
            "atc_chamado = '" + str(self.callId) + "'",
            "atc_data_hora_termino IS NULL"
            ]
        
        queryList.append(myDBConnection.updateQuery(tables, fieldsAndValues, conditions, queryBuild=True))
      

        ###################################################
        # -->After that insert the action in the database #
        ###################################################
        table = "acoes_chamados"

        fields = [
            "ach_descricao",
            "ach_detrator"
            ]

        values = [
            ("UPPER('" + actionTaked + "')", "'" + str(detractorId)  + "'")
            ]

        queryList.append(myDBConnection.insertQuery(table, fields, values, queryBuild=True))
            
        ######################################
        # --> And finally inserts the action #
        ######################################

        tables = ["chamados"]

        fieldsAndValues = [
            ("cha_status", "3"),
            ("cha_data_hora_termino", "NOW()"),
            ("cha_acao", "LAST_INSERT_ID()")
            ]
            
        conditions = [
            "cha_id = '" + str(self.callId) + "'"
            ]
            

        queryList.append(myDBConnection.updateQuery(tables, fieldsAndValues, conditions, queryBuild=True))

        querySuccess = myDBConnection.aglutinatedQuery(queryList)
 
        if not querySuccess:
           raise DatabaseConnectionError("Durante o fechamento dos chamados.",
                                         "Falha ao tentar executar o fechamento do chamado.\nContate o administrador do sistema.")
           return

###############################################################################################################################################################################

    def getActionTaken(self):
        """
        This method returns the detractor and the action taken to solve the call, if the call was answered
        """

        # --> DATABASE CONNECTION OBJECT #
        myDBConnection = DBConnection()
        
        if self.statusId != 3:
            return ("", "")
        
        fields = [
                "dtr_descricao",
                "ach_descricao"
                ]
            
        tables = [
            ("chamados", "", ""),            
            ("acoes_chamados", "LEFT", "cha_acao = ach_id"),
            ("detratores", "LEFT", "ach_detrator = dtr_id")
            ]

        where = [
             "cha_id = '" + str(self.callId)+ "' "
             ]

        querySuccess, queryResult = myDBConnection.selectQuery(fields, tables, where)
        if not querySuccess:
           raise DatabaseConnectionError("Durante o carregamento da ação tomada no relatório detalhado do chamado.",
                                         "Falha ao tentar executar a consulta no banco.\nContate o administrador do sistema.")
           return

        return queryResult[0]
        


###############################################################################################################################################################################
###############################################################################################################################################################################
###############################################################################################################################################################################
 

class CallContainer(object):
    """
    This class holds a set of Calls 
    """
    def __init__(self):
        self.__calls = []
        self.__callsFromId = {}
      
    
###############################################################################################################################################################################

    def verifyAnyCallBeingAnswered(self, idLoggedUser):
        """
        This method verifies if there is any call being answered by the user that opened the system
        """

        # --> DATABASE CONNECTION OBJECT #
        myDBConnection = DBConnection()
        
        fields = ["cha_id"]

        tables = [
            ("chamados", "", ""),
            ("atendimentos_chamados", "LEFT", "cha_id = atc_chamado")
            ]

        where = [
             "cha_status = 2",
             "atc_colaborador = '" + str(idLoggedUser) + "'",
             "atc_data_hora_termino IS NULL"
             ]

        querySuccess, queryResult = myDBConnection.selectQuery(fields, tables, where)

        if not querySuccess:
            raise DatabaseConnectionError("Verificação se já está respondendo chamado em Listagem de Chamados",
                                          "Falha ao tentar executar a consulta no banco.\nContate o administrador do sistema.")
            return

                       
        if len(queryResult) > 0:        
            return self.callFromId(queryResult[0][0])
        else:
            return None


###############################################################################################################################################################################
        
    def loadCalls(self):
        """
        This method starts the connection with the database and populates
        the list of calls
        """

        # --> DATABASE CONNECTION OBJECT #
        myDBConnection = DBConnection()

        # --> clear the data structure to load the new information about the calls #

        fields = [
                "cha_id",
                "cha_operador",
                "TIMESTAMPDIFF(MINUTE, cha_data_hora_abertura, NOW()) AS duracao_total",
                "IF(cha_status > 1, TIMESTAMPDIFF(MINUTE, cha_data_hora_atendimento, NOW()), 0) AS duracao_atendimento",
                "cha_tipo",
                "tch_descricao",
                "cli_nome",
                "pro_nome",
                "cha_DT",
                "cha_status",
                "stc_descricao",
                "atc_colaborador",
                "col_nome",
                "cha_descricao",
                "cha_plano",
                "cha_data_hora_abertura",
                "cha_data_hora_atendimento",
                "cha_data_hora_termino"
                ]

        tables = [
            ("chamados", "", ""),
            ("atendimentos_chamados", "LEFT", "cha_id = atc_chamado"),
            ("clientes", "LEFT", "cha_cliente = cli_id"),
            ("produtos", "LEFT", "cha_produto = pro_id"),
            ("colaboradores", "LEFT", "col_id = atc_colaborador"),
            ("tipos_chamado", "LEFT", "cha_tipo = tch_id"),
            ("status_chamado", "LEFT", "cha_status = stc_id")
            ]

        where = [
            "(cha_status = 1 OR cha_status = 2)",
            "atc_data_hora_termino IS NULL"
            ]

        groupby = ["cha_id"]

        orderby = [
            "cha_status DESC",
            "duracao_total DESC",
            "duracao_atendimento DESC"
           ]


        querySuccess, queryResult = myDBConnection.selectQuery(fields, tables, where, groupby, None, orderby)

        if not querySuccess:
            raise DatabaseConnectionError("Seleção dos Chamados na tela de Listagem de Chamados",
                                          "Falha ao tentar executar a consulta no banco.\nContate o administrador do sistema.")
            return
            
        self.clear()

        # --> populates the data structure #
        
        for (
            callId,
            creator,
            totalDuration,
            answerDuration,
            typeId,
            callType,
            client,
            product,
            device,
            statusId,
            status,
            supportId,
            support,
            description,
            inPlan,
            openingDate,
            answeringDate,
            endDate
            ) in queryResult:

            call = Call(callId, creator, totalDuration, answerDuration, typeId,
                        callType, client, product, device, statusId, status, supportId, support, description, inPlan, openingDate, answeringDate, endDate)
            self.add(call)


###############################################################################################################################################################################


    def loadCallsReport(self, filtering=False, filteringFields=None):
        """
        This method is used to load the calls in the calls report screen
        """

        # --> DATABASE CONNECTION OBJECT #
        myDBConnection = DBConnection()
        
        fields = [
                "cha_id",
                "cha_operador",
                "IF(cha_status < 3, TIMESTAMPDIFF(MINUTE, cha_data_hora_abertura, NOW()), TIMESTAMPDIFF(MINUTE, cha_data_hora_abertura, cha_data_hora_termino)) AS duracao_total",
                "IF(cha_status = 1, 0, IF(cha_status = 2, TIMESTAMPDIFF(MINUTE, cha_data_hora_atendimento, NOW()), TIMESTAMPDIFF(MINUTE, cha_data_hora_atendimento, cha_data_hora_termino))) AS duracao_atendimento",
                "cha_tipo",
                "tch_descricao",
                "cli_nome",
                "pro_nome",
                "cha_DT",
                "cha_status",
                "stc_descricao",
                "atc_colaborador",
                "col_nome",
                "cha_descricao",
                "cha_plano",
                "cha_data_hora_abertura",
                "cha_data_hora_atendimento",
                "cha_data_hora_termino",
                "dtr_descricao",
                ]
            
        tables = [
            ("chamados", "", ""),
            ("(SELECT atc1.* FROM atendimentos_chamados atc1 JOIN (SELECT atc_chamado, MAX(atc_data_hora_inicio) atc_inicio FROM atendimentos_chamados GROUP BY atc_chamado) atc2 ON atc1.atc_chamado = atc2.atc_chamado AND atc1.atc_data_hora_inicio = atc2.atc_inicio) atc", "LEFT", "atc.atc_chamado = cha_id"),
            ("clientes", "LEFT", "cha_cliente = cli_id"),
            ("produtos", "LEFT", "cha_produto = pro_id"),
            ("colaboradores", "LEFT", "col_id = atc_colaborador"),
            ("tipos_chamado", "LEFT", "cha_tipo = tch_id"),
            ("status_chamado", "LEFT", "cha_status = stc_id"),
            ("acoes_chamados", "LEFT", "cha_acao = ach_id"),
            ("detratores", "LEFT", "ach_detrator = dtr_id")
            ]


            
        
        if not filtering:
            where = [
                "DATE(cha_data_hora_abertura) >=  CURDATE() - INTERVAL 1 DAY",
                "DATE(cha_data_hora_abertura) <=  CURDATE()"
                ]
            limit = ["1000"]
        else:
            limit = None
            where = filteringFields

            
        groupby = ["cha_id"]

        orderby = [
            "cha_status ASC",
            "cha_data_hora_abertura DESC",
            "duracao_total DESC",
            "duracao_atendimento DESC",
            "atc_data_hora_inicio DESC"
           ]
        

        querySuccess, queryResult = myDBConnection.selectQuery(fields, tables, where, groupby, None, orderby, limit)
    
        if not querySuccess:
            raise DatabaseConnectionError("Seleção dos Chamados na tela de Listagem de Chamados",
                                          "Falha ao tentar executar a consulta no banco.\nContate o administrador do sistema.")
            return

        self.clear()
        
        # --> Populates the data structure #
        for (
            callId,
            creator,
            totalDuration,
            answerDuration,
            typeId,
            callType,
            client,
            product,
            device,
            statusId,
            status,
            supportId,
            support,
            description,
            inPlan,
            openingDate,
            answeringDate,
            endDate,
            detractor
            ) in queryResult:

            call = Call(callId, creator, totalDuration, answerDuration, typeId,
                        callType, client, product, device, statusId, status, supportId, support, description, inPlan, openingDate, answeringDate, endDate, detractor)
            self.add(call)


###############################################################################################################################################################################


    def add(self, call):
        """
        This method adds a new call to the call list. Returns True if success, False if fail
        """
        if int(call.callId) in self.__callsFromId.keys():
            return False

        self.__calls.append([int(call.callId), call])
        self.__callsFromId[int(call.callId)] = call
        return True

###############################################################################################################################################################################

    def clear(self):
        self.__calls = []
        self.__callsFromId = {}

###############################################################################################################################################################################
  
    def callFromId(self, callId):
        return self.__callsFromId[int(callId)]

###############################################################################################################################################################################

    def __iter__(self):
        for call in iter(self.__calls):
            yield call[1]

###############################################################################################################################################################################

    def __len__(self):
        return len(self.__calls)


###############################################################################################################################################################################
###############################################################################################################################################################################
###############################################################################################################################################################################


class CallType(object):
    """
    This class holds the information about one call type
    """
    def __init__(self, callTypeId, callTypeDescription):
        self.callTypeId = callTypeId
        self.callTypeDescription = callTypeDescription


###############################################################################################################################################################################
###############################################################################################################################################################################
###############################################################################################################################################################################


class CallTypeContainer(object):
    """
    This class is responsible for loading and holding all the call types
    """
    def __init__(self):
        self.__callTypes = []
        self.__callTypesFromId = {}

        
###############################################################################################################################################################################

    def loadCallTypes(self):
        """
        This method is called to load the call types from database
        """

        # --> DATABASE CONNECTION OBJECT #
        myDBConnection = DBConnection()
        
        fields = [
            "tch_id",
            "tch_descricao"
            ]

        tables = [("tipos_chamado", "", "")]

        orderby = ["tch_descricao ASC"]

        querySuccess, queryResult = myDBConnection.selectQuery(fields, tables, None, None, None, orderby)

        if not querySuccess:
            raise DatabaseConnectionError("Durante o carregamento dos tipos de chamado.",
                                          "Falha ao tentar executar a consulta no banco.\nContate o administrador do sistema.")
            return

        self.clear()
        
        for callTypeId, callTypeDescription in queryResult:
            callType = CallType(callTypeId, callTypeDescription)
            self.add(callType)

###############################################################################################################################################################################

    def add(self, callType):
        """
        This method adds a new call type to the call list. Returns True if success, False if fail
        """
        if int(callType.callTypeId) in self.__callTypesFromId.keys():
            return False

        self.__callTypes.append([int(callType.callTypeId), callType])
        self.__callTypesFromId[int(callType.callTypeId)] = callType
        return True

###############################################################################################################################################################################

    def clear(self):
        self.__callTypes = []
        self.__callTypesFromId = {}

###############################################################################################################################################################################
  
    def detractorFromId(self, callTypeId):
        return self.__callTypesFromId[int(callTypeId)]

###############################################################################################################################################################################

    def __iter__(self):
        for callTypes in iter(self.__callTypes):
            yield callTypes[1]

###############################################################################################################################################################################

    def __len__(self):
        return len(self.__callTypes)


###############################################################################################################################################################################
###############################################################################################################################################################################
###############################################################################################################################################################################


class CallStatus(object):
    """
    This class holds the information about one call status
    """
    def __init__(self, callStatusId, callStatusDescription):
        self.callStatusId = callStatusId
        self.callStatusDescription = callStatusDescription


###############################################################################################################################################################################
###############################################################################################################################################################################
###############################################################################################################################################################################


class CallStatusContainer(object):
    """
    This class is responsible for loading and honding all the call types
    """
    def __init__(self):
        self.__callStatus = []
        self.__callStatusFromId = {}
        

###############################################################################################################################################################################

    def loadCallStatus(self):
        """
        This method is called to load the call types from database
        """

        # --> DATABASE CONNECTION OBJECT #
        myDBConnection = DBConnection()
        
        fields = [
            "stc_id",
            "stc_descricao"
            ]

        tables = [("status_chamado", "", "")]

        orderby = ["stc_descricao ASC"]

        querySuccess, queryResult = myDBConnection.selectQuery(fields, tables, None, None, None, orderby)

        if not querySuccess:
            raise DatabaseConnectionError("Durante o carregamento dos status de chamado.",
                                          "Falha ao tentar executar a consulta no banco.\nContate o administrador do sistema.")
            return

        self.clear()
        
        for callStatusId, callStatusDescription in queryResult:
            callStatus = CallStatus(callStatusId, callStatusDescription)
            self.add(callStatus)

###############################################################################################################################################################################

    def add(self, callStatus):
        """
        This method adds a new call to the call list. Returns True if success, False if fail
        """
        if int(callStatus.callStatusId) in self.__callStatusFromId.keys():
            return False

        self.__callStatus.append([int(callStatus.callStatusId), callStatus])
        self.__callStatusFromId[int(callStatus.callStatusId)] = callStatus
        return True

###############################################################################################################################################################################

    def clear(self):
        self.__callStatus = []
        self.__callStatusFromId = {}

###############################################################################################################################################################################
  
    def detractorFromId(self, callStatusId):
        return self.__callStatusFromId[int(callStatusId)]

###############################################################################################################################################################################

    def __iter__(self):
        for callTypes in iter(self.__callStatus):
            yield callTypes[1]

###############################################################################################################################################################################

    def __len__(self):
        return len(self.__callStatus)


###############################################################################################################################################################################
###############################################################################################################################################################################
###############################################################################################################################################################################


class CallsIndicatorsReport(object):
    """
    This class is used to return useful informations about
    the indicators of an specified date
    """

    def __init__(self):
        self.totalCalls = 0
        self.totalAnsweringMinutes = 0
        self.totalLateMinutes = 0
        self.totalProductionMinutes = 0

###############################################################################################################################################################################

    def searchIndicator(self, dateBegin, dateEnd):
        ####################################################################
        # --> Firstly counts how many calls there are in the date interval #
        ####################################################################

        # --> DATABASE CONNECTION OBJECT #
        myDBConnection = DBConnection()
            
        fields = ["COUNT(cha_id)"]
            
        tables = [
            ("chamados", "", ""),
            ("acoes_chamados", "LEFT", "cha_acao = ach_id"),
            ("detratores", "LEFT", "ach_detrator = dtr_id")
            ]

        where = [
            "cha_status = 3",
            "cha_plano = 1",
            "DATE(cha_data_hora_abertura) >= DATE('" + dateBegin + "')",
            "DATE(cha_data_hora_abertura) <= DATE('" + dateEnd + "')",
            "dtr_indicador > 0"
            ]

        querySuccess, queryResult = myDBConnection.selectQuery(fields, tables, where)

        if not querySuccess:
            raise DatabaseConnectionError("Durante a primeira busca pelos indicadores num intervalo específico.",
                                          "Falha ao tentar executar a consulta no banco.\nContate o administrador do sistema.")
            return

        self.totalCalls = int(queryResult[0][0])

        fields = [
            "SUM(IF(ind_minutos_diario IS NULL, 0, ind_minutos_diario))",
            "SUM(IF(ind_atendimento_diario IS NULL, 0, ind_atendimento_diario))",
            "SUM(IF(ind_atraso_diario IS NULL, 0, ind_atraso_diario))"
            ]

        tables = [
            ("indicadores", "", "")
            ]

        where = [
            "ind_data >= '" + dateBegin + "'",
            "ind_data <= '" + dateEnd + "'"
            ]
            
        querySuccess, queryResult = myDBConnection.selectQuery(fields, tables, where)

        if not querySuccess:
            raise DatabaseConnectionError("Durante a segunda busca pelos indicadores num intervalo específico.",
                                          "Falha ao tentar executar a consulta no banco.\nContate o administrador do sistema.")
            return

        if queryResult[0][0]:
            self.totalProductionMinutes = int(queryResult[0][0])
            self.formatedTotalProductionMinutes = "%02d:%02d" % divmod(queryResult[0][0], 60)
        else:
            self.totalProductionMinutes = 0
            self.formatedTotalProductionMinutes = "00:00"

        if queryResult[0][1]:        
            self.totalAnsweringMinutes = int(queryResult[0][1])
            self.formatedTotalAnsweringMinutes = "%02d:%02d" % divmod(queryResult[0][1], 60)
        else:
            self.totalAnsweringMinutes = 0
            self.formatedTotalAnsweringMinutes = "00:00"

        if queryResult[0][2]:
            self.totalLateMinutes = int(queryResult[0][2])
            self.formatedtotalLateMinutes = "%02d:%02d" % divmod(queryResult[0][2], 60)
        else:
            self.totalLateMinutes = 0
            self.formatedtotalLateMinutes = "00:00"

################################################################################################################################################################################
################################################################################################################################################################################
################################################################################################################################################################################


class CallIndicators(object):
    """
    This class holds the statistics about the calls.
    The statistics created are daily, weekly and monthly
    """
    def __init__(self):
        self.dailyTotalCalls = 0
        self.dailyAvgAnswering = "00:00"
        self.dailyAvgLate = "00:00"
        self.dailyUpTime = "0,00%"

        self.weeklyTotalCalls = 0
        self.weeklyAvgAnswering = "00:00"
        self.weeklyAvgLate = "00:00"
        self.weeklyUpTime = "0,00%"

        self.monthlyTotalCalls = 0
        self.monthlyAvgAnswering = "00:00"
        self.monthlyAvgLate = "00:00"
        self.monthlyUpTime = "0,00%"

        
################################################################################################################################################################################

    def loadIndicators(self):
        """
        This method has as objective to calculate and to format the indicators, returning the string formated to be inserted
        in the indicators fields
        """

        # --> DATABASE CONNECTION OBJECT #
        myDBConnection = DBConnection()

        ######################################
        # --> CALCULATE THE DAILY INDICATORS #
        ######################################

        # --> first calculate the daily total number of calls, the total call answering time and the total call late time #
        
        fields = [
                "COUNT(cha_id)",
                "SUM( "\
                "   IF( "\
                "       cha_data_hora_abertura < cha_data_hora_termino, "\
                "       TIMESTAMPDIFF( "\
                "           MINUTE, "\
                "           IF( "\
		"		cha_data_hora_abertura > cha_data_hora_atendimento, "\
		"		cha_data_hora_abertura, "\
		"		cha_data_hora_atendimento "\
                "           ), "\
                "           cha_data_hora_termino "\
                "       ), "\
                "       0 "\
                "   ) "\
                ") AS duracao_atendimento, "\
                "SUM(IF(cha_data_hora_abertura < cha_data_hora_atendimento, "\
                    "TIMESTAMPDIFF(MINUTE, cha_data_hora_abertura, cha_data_hora_atendimento), "\
                    "0)) AS duracao_atraso"
                ]
            
        tables = [
            ("chamados", "", ""),
            ("acoes_chamados", "LEFT", "cha_acao = ach_id"),
            ("detratores", "LEFT", "ach_detrator = dtr_id")
            ]

        where = [
            "cha_status = 3",
            "cha_plano = 1",
            "DATE(cha_data_hora_abertura) = DATE(NOW())",
            "dtr_indicador > 0"
            ]

       
        querySuccess, queryResult = myDBConnection.selectQuery(fields, tables, where)

        if not querySuccess:
            raise DatabaseConnectionError("Durante a primeira busca pelas estatísticas diárias dos indicadores.",
                                          "Falha ao tentar executar a consulta no banco.\nContate o administrador do sistema.")
            return

        if len(queryResult) > 0:
            totalCalls, totalAnsweringTime, totalLateTime = queryResult[0]
            if totalAnsweringTime is None:
                totalAnsweringTime = 0
            if totalLateTime is None:
                totalLateTime = 0
            
            if totalCalls > 0:
                avgAnswering = round(totalAnsweringTime / totalCalls)
                avgLate = round(totalLateTime / totalCalls)
                self.dailyTotalCalls = totalCalls
                self.dailyAvgAnswering = "%02d:%02d" % divmod(avgAnswering, 60)
                self.dailyAvgLate = "%02d:%02d" % divmod(avgLate, 60)

        # --> after that calculates the daily uptime #

        fields = ["SUM(pdp_total_horas * 60)"]
            
        tables = [("planos_de_producao", "", "")]

        where = ["pdp_data = DATE(NOW())"]

        querySuccess, queryResult = myDBConnection.selectQuery(fields, tables, where)

        if not querySuccess:
            raise DatabaseConnectionError("Durante a segunda busca pelas estatísticas diárias dos indicadores.",
                                          "Falha ao tentar executar a consulta no banco.\nContate o administrador do sistema.")
            return

         
        if len(queryResult) > 0:
            if queryResult[0][0]:
                uptime = 1 - ((int(totalAnsweringTime) + int(totalLateTime)) / int(queryResult[0][0]))
                self.dailyUpTime = "{:.2f}%".format(uptime*100)
            else:
                self.dailyUpTime = "100,00%"
        else:
            self.dailyUpTime = "0,00%"

        

        #######################################
        # --> CALCULATE THE WEEKLY INDICATORS #
        #######################################

        # --> first calculate the weekly total number of calls, the total call answering time and the total call late time #
        
        fields = [
                "COUNT(cha_id)",
                "SUM( "\
                "   IF( "\
                "       cha_data_hora_abertura < cha_data_hora_termino, "\
                "       TIMESTAMPDIFF( "\
                "           MINUTE, "\
                "           IF( "\
		"		cha_data_hora_abertura > cha_data_hora_atendimento, "\
		"		cha_data_hora_abertura, "\
		"		cha_data_hora_atendimento "\
                "           ), "\
                "           cha_data_hora_termino "\
                "       ), "\
                "       0 "\
                "   ) "\
                ") AS duracao_atendimento, "\
                "SUM(IF(cha_data_hora_abertura < cha_data_hora_atendimento, "\
                    "TIMESTAMPDIFF(MINUTE, cha_data_hora_abertura, cha_data_hora_atendimento), "\
                    "0)) AS duracao_atraso"
                ]
            
        tables = [
            ("chamados", "", ""),
            ("acoes_chamados", "LEFT", "cha_acao = ach_id"),
            ("detratores", "LEFT", "ach_detrator = dtr_id")
            ]

        where = [
            "cha_status = 3",            
            "cha_plano = 1",
            "WEEK(cha_data_hora_abertura) = WEEK(NOW())",
            "YEAR(cha_data_hora_abertura) = YEAR(NOW())",
            "dtr_indicador > 0"
            ]

        
        querySuccess, queryResult = myDBConnection.selectQuery(fields, tables, where)

        if not querySuccess:
            raise DatabaseConnectionError("Durante a primeira busca pelas estatísticas semanais dos indicadores.",
                                          "Falha ao tentar executar a consulta no banco.\nContate o administrador do sistema.")
            return
        
        if len(queryResult) > 0:
            totalCalls, totalAnsweringTime, totalLateTime = queryResult[0]
            if totalAnsweringTime is None:
                totalAnsweringTime = 0
            if totalLateTime is None:
                totalLateTime = 0

            if totalCalls > 0:
                avgAnswering = round(totalAnsweringTime / totalCalls)
                avgLate = round(totalLateTime / totalCalls)
                self.weeklyTotalCalls = totalCalls
                self.weeklyAvgAnswering = "%02d:%02d" % divmod(avgAnswering, 60)
                self.weeklyAvgLate = "%02d:%02d" % divmod(avgLate, 60)

        # --> after that calculates the weekly uptime #

        fields = ["SUM(pdp_total_horas * 60)"]
            
        tables = [("planos_de_producao", "", "")]

        where = [
            "WEEK(pdp_data) = WEEK(NOW())",            
            "YEAR(pdp_data) = YEAR(NOW())"
        ]

        querySuccess, queryResult = myDBConnection.selectQuery(fields, tables, where)

        if not querySuccess:
            raise DatabaseConnectionError("Durante a segunda busca pelas estatísticas semanais dos indicadores.",
                                          "Falha ao tentar executar a consulta no banco.\nContate o administrador do sistema.")
            return

        
        if len(queryResult) > 0:
            if queryResult[0][0]:
                uptime = 1 - ((int(totalAnsweringTime) + int(totalLateTime)) / int(queryResult[0][0]))
                self.weeklyUpTime = "{:.2f}%".format(uptime*100)
            else:
                self.weeklyUpTime = "100,00%"
        else:
            self.weeklyUpTime = "0,00%"
            

        ########################################
        # --> CALCULATE THE MONTHLY INDICATORS #
        ########################################

        # --> first calculate the monthly total number of calls, the total call answering time and the total call late time #

        fields = [
            "COUNT(cha_id)",
            "SUM( "\
                "   IF( "\
                "       cha_data_hora_abertura < cha_data_hora_termino, "\
                "       TIMESTAMPDIFF( "\
                "           MINUTE, "\
                "           IF( "\
		"		cha_data_hora_abertura > cha_data_hora_atendimento, "\
		"		cha_data_hora_abertura, "\
		"		cha_data_hora_atendimento "\
                "           ), "\
                "           cha_data_hora_termino "\
                "       ), "\
                "       0 "\
                "   ) "\
                ") AS duracao_atendimento, "\
            "SUM(IF(cha_data_hora_abertura < cha_data_hora_atendimento, "\
                "TIMESTAMPDIFF(MINUTE, cha_data_hora_abertura, cha_data_hora_atendimento), "\
                "0)) AS duracao_atraso"
            ]
            
        tables = [
            ("chamados", "", ""),
            ("acoes_chamados", "LEFT", "cha_acao = ach_id"),
            ("detratores", "LEFT", "ach_detrator = dtr_id")
            ]

        where = [
            "cha_status = 3",
            "cha_plano = 1",
            "MONTH(cha_data_hora_abertura) = MONTH(NOW())",
            "YEAR(cha_data_hora_abertura) = YEAR(NOW())",
            "dtr_indicador > 0"
            ]

               
        querySuccess, queryResult = myDBConnection.selectQuery(fields, tables, where)

        if not querySuccess:
            raise DatabaseConnectionError("Durante a primeira busca pelas estatísticas mensais dos indicadores.",
                                          "Falha ao tentar executar a consulta no banco.\nContate o administrador do sistema.")
            return
        
        if len(queryResult) > 0:
            totalCalls, totalAnsweringTime, totalLateTime = queryResult[0]
            if totalAnsweringTime is None:
                totalAnsweringTime = 0
            if totalLateTime is None:
                totalLateTime = 0

            if totalCalls > 0:
                avgAnswering = round(totalAnsweringTime / totalCalls)
                avgLate = round(totalLateTime / totalCalls)
                self.monthlyTotalCalls = totalCalls
                self.monthlyAvgAnswering = "%02d:%02d" % divmod(avgAnswering, 60)
                self.monthlyAvgLate = "%02d:%02d" % divmod(avgLate, 60)

        # --> after that calculates the weekly uptime #

        fields = ["SUM(pdp_total_horas * 60)"]
            
        tables = [("planos_de_producao", "", "")]

        where = [
            "MONTH(pdp_data) = MONTH(NOW())",
            "YEAR(pdp_data) = YEAR(NOW())"
        ]

        querySuccess, queryResult = myDBConnection.selectQuery(fields, tables, where)

        if not querySuccess:
            raise DatabaseConnectionError("Durante a segunda busca pelas estatísticas mensais dos indicadores.",
                                          "Falha ao tentar executar a consulta no banco.\nContate o administrador do sistema.")
            return

        
        if len(queryResult) > 0:
            if queryResult[0][0]:
                uptime = 1 - ((int(totalAnsweringTime) + int(totalLateTime)) / int(queryResult[0][0]))
                self.monthlyUpTime = "{:.2f}%".format(uptime*100)
            else:
                self.monthlyUpTime = "100,00%"
        else:
            self.monthlyUpTime = "0,00%"

################################################################################################################################################################################
################################################################################################################################################################################
################################################################################################################################################################################


class CallsWorksheetReport(object):
    """
    This class is responsible for building the excel file containing all the call
    data of the report
    """
    def __init__(self, reportPath):
        self.reportPath = reportPath
       
###############################################################################################################################################################################
    
    def generateReport(self, filtering=False, filteringFields=None):
        """
        This method is used to load the calls in the calls report screen
        """

        # --> DATABASE CONNECTION OBJECT #
        myDBConnection = DBConnection()
        
        fields = [
                "cha_id",
                "cha_operador",
                "IF(cha_status < 3, TIMESTAMPDIFF(MINUTE, cha_data_hora_abertura, NOW()), TIMESTAMPDIFF(MINUTE, cha_data_hora_abertura, cha_data_hora_termino)) AS duracao_total",
                "IF(cha_status >= 2, TIMESTAMPDIFF(MINUTE, cha_data_hora_abertura, cha_data_hora_atendimento), TIMESTAMPDIFF(MINUTE, cha_data_hora_abertura, NOW())) AS duracao_atraso",
                "IF(cha_status = 1, 0, IF(cha_status = 2, TIMESTAMPDIFF(MINUTE, cha_data_hora_atendimento, NOW()), TIMESTAMPDIFF(MINUTE, cha_data_hora_atendimento, cha_data_hora_termino))) AS duracao_atendimento",
                "cha_tipo",
                "tch_descricao",
                "cli_nome",
                "pro_nome",
                "cha_DT",
                "cha_status",
                "stc_descdricao",
                "atc_colaborador",
                "col_nome",
                "cha_descricao",
                "cha_plano",
                "DATE_FORMAT(cha_data_hora_abertura, '%d/%m/%Y %H:%i') as cha_data_hora_abertura",
                "cha_data_hora_atendimento",
                "cha_data_hora_termino",
                "dtr_descricao",
                "ach_descricao"
                ]
            
        tables = [
            ("chamados", "", ""),
            ("(SELECT atc1.* FROM atendimentos_chamados atc1 JOIN (SELECT atc_chamado, MAX(atc_data_hora_inicio) atc_inicio FROM atendimentos_chamados GROUP BY atc_chamado) atc2 ON atc1.atc_chamado = atc2.atc_chamado AND atc1.atc_data_hora_inicio = atc2.atc_inicio) atc", "LEFT", "atc.atc_chamado = cha_id"),
            ("clientes", "LEFT", "cha_cliente = cli_id"),
            ("produtos", "LEFT", "cha_produto = pro_id"),
            ("colaboradores", "LEFT", "col_id = atc_colaborador"),
            ("tipos_chamado", "LEFT", "cha_tipo = tch_id"),
            ("status_chamado", "LEFT", "cha_status = stc_id"),
            ("acoes_chamados", "LEFT", "cha_acao = ach_id"),
            ("detratores", "LEFT", "ach_detrator = dtr_id")
            ]


            
        
        if not filtering:
            where = [
                "DATE(cha_data_hora_abertura) >=  CURDATE() - INTERVAL 1 DAY",
                "DATE(cha_data_hora_abertura) <=  CURDATE()"
                ]
            limit = ["1000"]
        else:
            limit = None
            where = filteringFields

            
        groupby = ["cha_id"]

        orderby = [
            "cha_status ASC",
            "cha_data_hora_abertura DESC",
            "duracao_total DESC",
            "duracao_atendimento DESC",
            "atc_data_hora_inicio DESC"
           ]
        

        querySuccess, queryResult = myDBConnection.selectQuery(fields, tables, where, groupby, None, orderby, limit)
    
        if not querySuccess:
            raise DatabaseConnectionError("Seleção dos Chamados na tela de Listagem de Chamados",
                                          "Falha ao tentar executar a consulta no banco.\nContate o administrador do sistema.")
            return

        wb = Workbook()
        destiny = self.reportPath + "/relatorio.xlsx"
        workSheet = wb.active
        workSheet.title = "Relatório"

        workSheet.cell(column=1, row=1, value="DATA DE ABERTURA")
        workSheet.cell(column=2, row=1, value="CHAMADO NO PLANO")
        workSheet.cell(column=3, row=1, value="TIPO DE CHAMADO")
        workSheet.cell(column=4, row=1, value="DURAÇÃO TOTAL DO CHAMADO")
        workSheet.cell(column=5, row=1, value="TEMPO TOTAL DE ATRASO")
        workSheet.cell(column=6, row=1, value="TEMPO TOTAL DE ATENDIMENTO")
        workSheet.cell(column=7, row=1, value="STATUS DO CHAMADO")
        workSheet.cell(column=8, row=1, value="CRIADOR DO CHAMADO")
        workSheet.cell(column=9, row=1, value="CLIENTE")
        workSheet.cell(column=10, row=1, value="PRODUTO")
        workSheet.cell(column=11, row=1, value="SUPORTE RESPONSÁVEL")
        workSheet.cell(column=12, row=1, value="DETRATOR")
        workSheet.cell(column=13, row=1, value="AÇÃO REALIZADA")

        workSheet.column_dimensions['A'].width = 20
        workSheet.column_dimensions['B'].width = 20
        workSheet.column_dimensions['C'].width = 20
        workSheet.column_dimensions['D'].width = 30
        workSheet.column_dimensions['E'].width = 23
        workSheet.column_dimensions['F'].width = 30
        workSheet.column_dimensions['G'].width = 22
        workSheet.column_dimensions['H'].width = 25
        workSheet.column_dimensions['I'].width = 25
        workSheet.column_dimensions['J'].width = 25
        workSheet.column_dimensions['K'].width = 30
        workSheet.column_dimensions['L'].width = 50
        workSheet.column_dimensions['M'].width = 100
        

        
        row = 2
        for (
            callId,
            creator,
            totalDuration,
            lateDuration,
            answerDuration,
            typeId,
            callType,
            client,
            product,
            device,
            statusId,
            status,
            supportId,
            support,
            description,
            inPlan,
            openingDate,
            answeringDate,
            endDate,
            detractor,
            action
            )in queryResult:
                if totalDuration < 0:
                    totalDuration = abs(int(totalDuration))
                    formatedTotalDuration = "-%02d:%02d" % (totalDuration//60, totalDuration%60)
                else:
                    formatedTotalDuration = "%02d:%02d" % (totalDuration//60, totalDuration%60)

                if lateDuration < 0:
                    lateDuration = abs(int(lateDuration))
                    formatedLateDuration = "-%02d:%02d" % (lateDuration//60, lateDuration%60)
                else:
                    formatedLateDuration = "%02d:%02d" % (lateDuration//60, lateDuration%60)

                if answerDuration < 0:
                    answerDuration = abs(int(answerDuration))
                    formatedAnswerDuration = "-%02d:%02d" % (answerDuration//60, answerDuration%60)
                else:
                    formatedAnwerDuration = "%02d:%02d" % (answerDuration//60, answerDuration%60)

                    
                workSheet.cell(column=1, row=row, value=openingDate)
                workSheet.cell(column=2, row=row, value=inPlan)
                workSheet.cell(column=3, row=row, value=callType)
                workSheet.cell(column=4, row=row, value=formatedTotalDuration)
                workSheet.cell(column=5, row=row, value=formatedLateDuration)
                workSheet.cell(column=6, row=row, value=formatedAnwerDuration)
                workSheet.cell(column=7, row=row, value=status)
                workSheet.cell(column=8, row=row, value=creator)
                workSheet.cell(column=9, row=row, value=client)
                workSheet.cell(column=10, row=row, value=product)
                workSheet.cell(column=11, row=row, value=support)
                workSheet.cell(column=12, row=row, value=detractor)
                workSheet.cell(column=13, row=row, value=action)
                

                row = row + 1
                

        wb.save(filename = destiny)

################################################################################################################################################################################
################################################################################################################################################################################
################################################################################################################################################################################

class CallCreator(object):
    """
    This class is responsible for creating a object that will insert a
    engineering call into the system
    """
    def __init__(self, clientId, productId, deviceId, callDescription, callUser):
        self.clientId = clientId
        self.productId = productId
        self.deviceId = deviceId
        self.callDescription = callDescription
        self.callUser = callUser


    def createEngineerCall(self):
        """
        This method inserts the object info inside the call table
        """
        # --> DATABASE CONNECTION OBJECT #
        myDBConnection = DBConnection()
        
        table = "chamados"

        fields = [
            "cha_tipo",
            "cha_cliente",
            "cha_produto",
            "cha_DT",
            "cha_descricao",
            "cha_status",
            "cha_data_hora_abertura",
            "cha_operador",
            "cha_plano"
            ]

        values = [
            ("'5'",
             "'" + str(self.clientId) + "'",
             "'" + str(self.productId) + "'",
             "'" + str(self.deviceId) + "'",
             "UPPER('" + str(self.callDescription) + "')",
             "'1'",
             "NOW()",
             "'" + str(self.callUser) + "'",
             "'-1'")]

        inserted, insertedId = myDBConnection.insertQuery(table, fields, values)

        if not inserted:
            raise DatabaseConnectionError("Durante a criação de um chamado de engenharia.",
                                          "Falha ao tentar executar a inserção no banco.\nContate o administrador do sistema.")
            return 




